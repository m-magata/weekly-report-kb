"""
Excelパーサー: 週報Excelファイルを読み込みDBモデル用データに変換する。

対応シート:
  - 「売上進捗表」: 日別売上・客数・天候（週ブロック10行単位）
  - 「週報①〜⑥」: 店長報告テキスト（行17以降 ＜営業報告＞ セクション）

実際のExcel構造 (041 販売部週報 26-03.xlsx 調査済み):
  売上進捗表:
    行1  : (None,None,None,None, 年, '年', 月, '月度...')
    行2  : (None,None,..., 店舗名, None, '店', None)
    週ブロック (各10行):
      row+0: (週名称, '日付', d1, d2, d3, d4, d5, d6, d7, None, 週名称, ...)
      row+1: (None, '天候', w1, w2, w3, w4, w5, w6, w7, ...)
      row+2: (None, '売上', s1, s2, s3, s4, s5, s6, s7, ...)
      row+6: (None, '客数', c1, c2, c3, c4, c5, c6, c7, ...)
  週報① ~ 週報⑥:
    行17 : '＜営業報告＞'
    行18+: 報告テキスト本文
"""

from __future__ import annotations

import re
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path


class SkipFileError(Exception):
    """パース対象外ファイル（フォーマットファイル等）を示す例外。"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import xlrd as _xlrd


# ---------------------------------------------------------------------------
# xlrd ラッパー（.xls ファイル用 / openpyxl ライクなインターフェース）
# ---------------------------------------------------------------------------

class _XlrdCellWrapper:
    """xlrd セルをopenpyxlライクにラップ（書式情報なし）。"""
    __slots__ = ("value", "fill")

    def __init__(self, raw_value):
        # xlrd の空セル（''）を None に統一
        self.value = None if raw_value == "" else raw_value
        self.fill = None  # xls は書式未取得のため黒背景判定は常に False


class _XlrdSheetWrapper:
    """xlrd シートをopenpyxlライクにラップ。"""

    def __init__(self, xlrd_sheet):
        self._sheet = xlrd_sheet

    def iter_rows(self, min_row: int = 1, max_row: int | None = None, values_only: bool = True):
        nrows = self._sheet.nrows
        end_row = nrows if max_row is None else min(max_row, nrows)
        for r in range(min_row - 1, end_row):
            raw = self._sheet.row_values(r)
            if values_only:
                yield tuple(None if v == "" else v for v in raw)
            else:
                yield [_XlrdCellWrapper(v) for v in raw]


class _XlrdWorkbookWrapper:
    """xlrd Workbook をopenpyxlライクにラップ。"""

    def __init__(self, xlrd_book):
        self._book = xlrd_book

    @property
    def sheetnames(self) -> list[str]:
        return self._book.sheet_names()

    def __contains__(self, name: str) -> bool:
        return name in self._book.sheet_names()

    def __getitem__(self, name: str) -> _XlrdSheetWrapper:
        return _XlrdSheetWrapper(self._book.sheet_by_name(name))


# 「週報①」〜「週報⑥」に対応
REPORT_SHEET_PATTERN = re.compile(r"^週報[①②③④⑤⑥]$")
SHEET_INDEX_MAP = {"①": 1, "②": 2, "③": 3, "④": 4, "⑤": 5, "⑥": 6}

# 売上進捗表の週ブロック内の行オフセット（0始まり）
ROW_DATE = 0
ROW_WEATHER = 1
ROW_SALES = 2
ROW_SALES_RATIO = 5      # 予算比（実売÷日割予算）
ROW_CUSTOMERS = 6
ROW_CUSTOMERS_PREV = 7   # 昨客（昨年同週客数）

# 日付データが入る列インデックス（0始まり、C〜I列 = 2〜8）
DATA_COL_START = 2
DATA_COL_END = 9  # exclusive (7日分)

# 来期コメント欄などに混入する署名セルを除去
SIGNATURE_WORDS = frozenset({"社長", "部長", "エリア長", "次長", "課長", "副部長"})


@dataclass
class DailySalesRecord:
    date: date
    sales_amount: float | None  # 千円単位
    customer_count: int | None
    weather: str | None
    sales_budget_ratio: float | None = None      # 売上予算比（実売÷日割予算）
    customer_count_prev_year: int | None = None  # 昨年同週客数


@dataclass
class ReportTextRecord:
    sheet_index: int  # 1〜6
    sheet_name: str
    content: str


@dataclass
class ParsedReport:
    store_name: str
    manager_name: str | None
    week_start: date
    week_end: date
    source_filename: str
    submitter_role: str = "店長"  # '店長' or '副店長'
    report_year: int | None = None
    report_month: int | None = None
    daily_sales: list[DailySalesRecord] = field(default_factory=list)
    report_texts: list[ReportTextRecord] = field(default_factory=list)


def parse_excel(filepath: str | Path) -> ParsedReport:
    """Excelファイルを読み込み ParsedReport を返す。

    Raises:
        SkipFileError: store_code='0000' のフォーマットファイルの場合。
    """
    path = Path(filepath)

    # Excel 一時ファイル（~$ で始まる）はスキップ
    if path.name.startswith("~$"):
        raise SkipFileError(f"Excel一時ファイルのためスキップします: {path.name}")

    # store_code='0000' のファイルはフォーマットファイルのためスキップ
    m = re.match(r"^(\d+)", path.name)
    if m and m.group(1).zfill(4) == "0000":
        raise SkipFileError(f"フォーマットファイルのためスキップします: {path.name}")

    is_fuku = _is_fuku_tencho(path.name)

    suffix = path.suffix.lower()
    if suffix == ".xls":
        wb = _XlrdWorkbookWrapper(_xlrd.open_workbook(str(path)))
    else:
        wb = openpyxl.load_workbook(path, data_only=True)

    if "売上進捗表" not in wb.sheetnames:
        raise SkipFileError(f"週報フォーマット外のファイルのためスキップします: {path.name}")
    store_name, year, month = _extract_header_info(wb["売上進捗表"])

    # 副店長ファイルは店舗名末尾に「副」を付加
    if is_fuku:
        store_name = store_name.rstrip("店") + "店副"

    daily_sales = _parse_sales_sheet(wb["売上進捗表"], year, month)
    report_texts = _parse_report_sheets(wb)
    week_start, week_end = _infer_week_range(daily_sales)

    ym = _extract_report_ym_from_filename(path.name)

    return ParsedReport(
        store_name=store_name,
        manager_name=None,
        week_start=week_start,
        week_end=week_end,
        source_filename=path.name,
        submitter_role="副店長" if is_fuku else "店長",
        report_year=ym[0] if ym else None,
        report_month=ym[1] if ym else None,
        daily_sales=daily_sales,
        report_texts=report_texts,
    )


# ---------------------------------------------------------------------------
# 売上進捗表パース
# ---------------------------------------------------------------------------

def _is_fuku_tencho(filename: str) -> bool:
    """ファイル名から副店長提出レポートかどうかを判定する。

    対応パターン:
      - '副店長' を含む場合（例: 052副店長販売部週報 26-03.xlsx）
      - ファイル名（拡張子なし）が '副' で終わる場合（例: 069販売部週報26-03副.xlsx）
    """
    stem = Path(filename).stem
    return "副店長" in filename or stem.endswith("副")


def _extract_header_info(ws: Worksheet) -> tuple[str, int, int]:
    """
    売上進捗表の先頭2行から年・月・店舗名を抽出。

    行1: (None,None,None,None, 2026, '年', 3, '月度...')
    行2: (None,None,..., 'ＮＴ南', None, '店', None)
    """
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))

    # 年・月: 行1のセルを巡回
    year, month = date.today().year, date.today().month
    row1 = rows[0]
    for i, val in enumerate(row1):
        if val == "年" and i > 0 and isinstance(row1[i - 1], (int, float)):
            year = int(row1[i - 1])
        if isinstance(val, str) and "月" in val and i > 0 and isinstance(row1[i - 1], (int, float)):
            month = int(row1[i - 1])

    # 店舗名: 行2の「店」の前後を探す（間にNoneが入る場合を考慮）
    # 例: [..., 'ＮＴ南', None, '店', ...]
    # 候補に既に「店」が含まれる場合は重複しないよう末尾処理する
    store_name = "不明"
    row2 = rows[1]
    for i, val in enumerate(row2):
        if val == "店" and i > 1:
            # i-1 または i-2 に店舗名候補があれば採用
            for offset in (1, 2):
                candidate = row2[i - offset] if i >= offset else None
                if candidate and str(candidate).strip():
                    raw = str(candidate).strip()
                    store_name = raw if raw.endswith("店") else raw + "店"
                    break
            if store_name != "不明":
                break

    return store_name, year, month


def _parse_sales_sheet(
    ws: Worksheet, year: int, month: int
) -> list[DailySalesRecord]:
    """
    売上進捗表の週ブロックを解析し日別レコードリストを返す。

    週ブロック検出条件: 列A(index 0)が文字列かつ列B(index 1)=='日付'
    """
    all_rows = list(ws.iter_rows(values_only=True))
    records: list[DailySalesRecord] = []

    # 月またぎを追跡するため、前回処理した日を保持
    prev_day: int | None = None
    current_month = month - 1 if month > 1 else 12
    current_year = year if month > 1 else year - 1

    for i, row in enumerate(all_rows):
        # 週ブロック先頭行の検出
        if not (_cell_str(row, 0) and _cell_str(row, 1) == "日付"):
            continue

        # 日付行
        date_row = row
        # 各補助行が範囲内にあるか確認（最大オフセット ROW_CUSTOMERS_PREV=7 で判定）
        if i + ROW_CUSTOMERS_PREV >= len(all_rows):
            continue
        weather_row = all_rows[i + ROW_WEATHER]
        sales_row = all_rows[i + ROW_SALES]
        ratio_row = all_rows[i + ROW_SALES_RATIO]
        customer_row = all_rows[i + ROW_CUSTOMERS]
        prev_cust_row = all_rows[i + ROW_CUSTOMERS_PREV]

        for col in range(DATA_COL_START, DATA_COL_END):
            day_val = date_row[col] if col < len(date_row) else None
            if not isinstance(day_val, (int, float)) or day_val <= 0:
                continue
            day = int(day_val)

            # 月またぎ検出: 日が前の日より小さくなったら翌月へ
            if prev_day is not None and day < prev_day:
                if current_month == 12:
                    current_month = 1
                    current_year += 1
                else:
                    current_month += 1
            prev_day = day

            try:
                record_date = date(current_year, current_month, day)
            except ValueError:
                continue

            sales = _to_float(sales_row[col] if col < len(sales_row) else None)
            customers = _to_int(customer_row[col] if col < len(customer_row) else None)
            weather_raw = weather_row[col] if col < len(weather_row) else None
            weather = str(weather_raw).strip() if weather_raw else None
            sales_ratio = _to_float(ratio_row[col] if col < len(ratio_row) else None)
            prev_customers = _to_int(prev_cust_row[col] if col < len(prev_cust_row) else None)

            # 売上も客数もNullの行はスキップ（未来週）
            if sales is None and customers is None:
                continue

            records.append(DailySalesRecord(
                date=record_date,
                sales_amount=sales,
                customer_count=customers,
                weather=weather,
                sales_budget_ratio=sales_ratio,
                customer_count_prev_year=prev_customers,
            ))

    return records


# ---------------------------------------------------------------------------
# 週報テキストパース
# ---------------------------------------------------------------------------

def _parse_report_sheets(wb: openpyxl.Workbook) -> list[ReportTextRecord]:
    """週報①〜⑦シートから営業報告テキストを抽出する。

    values_only=False でセルオブジェクトを取得し、黒/ダーク系の背景色セルの
    テキスト先頭に「★」プレフィックスを付与する（フロント側で強調表示に使用）。
    """
    records: list[ReportTextRecord] = []

    for sheet_name in wb.sheetnames:
        if not REPORT_SHEET_PATTERN.match(sheet_name):
            continue

        last_char = sheet_name[-1]
        sheet_index = SHEET_INDEX_MAP.get(last_char, 0)
        ws = wb[sheet_name]

        text_lines: list[str] = []
        in_report_section = False

        for row in ws.iter_rows(values_only=False):
            row_texts: list[str] = []
            found_marker = False

            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text:
                    continue
                # 署名欄の単独セル（「社長」「部長」等）を除去
                if text in SIGNATURE_WORDS:
                    continue
                # 山括弧付きの「＜営業報告＞」を正確に検出
                if not in_report_section and "＜" in text and "営業報告" in text:
                    found_marker = True
                    continue
                if in_report_section:
                    # セル単位で黒背景を判定し、そのセルの行頭にのみ★を付与
                    if _cell_has_dark_fill(cell):
                        row_texts.append("★" + text)
                    else:
                        row_texts.append(text)

            if found_marker:
                in_report_section = True

            if in_report_section and not found_marker and row_texts:
                text_lines.extend(row_texts)

        content = "\n".join(text_lines).strip()
        if content:
            records.append(ReportTextRecord(
                sheet_index=sheet_index,
                sheet_name=sheet_name,
                content=content,
            ))

    records.sort(key=lambda r: r.sheet_index)
    return records


def _cell_has_dark_fill(cell) -> bool:
    """セルの背景色が黒/ダーク系かを判定する。

    実測値:
      通常セル : fill_type='solid', fgColor.type='indexed', idx=41 (白)
      黒背景セル: fill_type='solid', fgColor.type='theme',   theme=1, tint=0.0 (黒)
    """
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    color = fill.fgColor
    if color is None:
        return False
    if color.type == "rgb":
        rgb = color.rgb or ""
        if len(rgb) == 8 and rgb.upper() not in ("00000000", "FFFFFFFF"):
            r = int(rgb[2:4], 16)
            g = int(rgb[4:6], 16)
            b = int(rgb[6:8], 16)
            if (0.299 * r + 0.587 * g + 0.114 * b) / 255 < 0.35:
                return True
    elif color.type == "theme":
        # theme=1 (Text 1) かつ tint が 0 以下（明化なし）= 黒相当
        try:
            if color.theme == 1 and color.tint <= 0.0:
                return True
        except Exception:
            pass
    return False


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def _extract_report_ym_from_filename(filename: str) -> tuple[int, int] | None:
    """ファイル名から年月を抽出する。

    対応パターン（各種ダッシュ記号を許容）:
      '065 販売部週報 26-03.xlsx' → (2026, 3)
      '073販売部週報25-10.xlsx'   → (2025, 10)
      '006副店長販売部週報25ー1.xls' → (2025, 1)

    ファイル名に含まれる 2桁年 + ダッシュ + 1〜2桁月 パターンを全て検索し、
    最後のマッチを採用する（先頭の店舗コード数字との混同を避けるため）。
    """
    stem = Path(filename).stem
    # 各種ダッシュ: ハイフン / 全角マイナス / 長音記号（ーｰ）
    pattern = re.compile(r'(\d{2})\s*[－ｰーｰ\-]\s*(\d{1,2})')
    matches = pattern.findall(stem)
    if not matches:
        return None
    yy, mm = matches[-1]
    year = 2000 + int(yy)
    month = int(mm)
    if not (1 <= month <= 12):
        return None
    return year, month


def _infer_week_range(records: list[DailySalesRecord]) -> tuple[date, date]:
    if not records:
        today = date.today()
        return today, today
    dates = [r.date for r in records]
    return min(dates), max(dates)


def _cell_str(row: tuple, index: int) -> str | None:
    if index >= len(row):
        return None
    val = row[index]
    if val is None:
        return None
    return str(val).strip() or None


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned or cleaned in ("#DIV/0!", "-", ""):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_int(value: object) -> int | None:
    f = _to_float(value)
    return int(f) if f is not None else None
